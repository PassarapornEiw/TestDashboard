"""
Excel Library for Robot Framework
Handles reading and writing Excel files for test data management
"""
from typing import Any, Dict, List, Optional
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet


class ExcelLibrary:
    """Library for Excel file operations"""

    ROBOT_LIBRARY_SCOPE = 'GLOBAL'
    ROBOT_LIBRARY_VERSION = '1.0'

    ALTERNATIVE_SHEET_NAMES = {"TestCases": "Main", "Main": "TestCases"}

    def __init__(self):
        self.workbook: Optional[Workbook] = None
        self.worksheet: Optional[Worksheet] = None
        self.file_path: Optional[Path] = None

    def load_excel_file(self, file_path: str) -> Workbook:
        """Load Excel file"""
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        self.workbook = load_workbook(file_path)
        return self.workbook

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """Get worksheet by name"""
        if not self.workbook:
            raise ValueError("No workbook loaded. Call load_excel_file first.")

        # Try exact match first
        if sheet_name in self.workbook.sheetnames:
            self.worksheet = self.workbook[sheet_name]
            return self.worksheet

        # Try alternative names
        if sheet_name in self.ALTERNATIVE_SHEET_NAMES:
            alt_name = self.ALTERNATIVE_SHEET_NAMES[sheet_name]
            if alt_name in self.workbook.sheetnames:
                self.worksheet = self.workbook[alt_name]
                return self.worksheet

        available_sheets = ", ".join(self.workbook.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {available_sheets}")

    def _get_headers(self, sheet_name: str) -> tuple[Dict[str, int], int]:
        """Get headers from sheet and return headers dict and header row number"""
        if not self.worksheet or self.worksheet.title != sheet_name:
            self.get_sheet(sheet_name)

        headers: Dict[str, int] = {}
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=1, max_row=10, values_only=False), 1):
            values = [cell.value for cell in row]
            if 'TestCaseNo' in values:
                for col_idx, cell in enumerate(row, 1):
                    if cell.value:
                        headers[cell.value] = col_idx
                return headers, row_idx

        raise ValueError("Header row with 'TestCaseNo' not found")

    def get_test_case_data(self, test_case_no: str, sheet_name: str = "TestCases") -> Dict[str, Any]:
        """Get test case data by TestCaseNo"""
        headers, header_row = self._get_headers(sheet_name)
        test_case_col = headers.get('TestCaseNo')

        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=header_row + 1, values_only=False), header_row + 1):
            if row[test_case_col - 1].value == test_case_no:
                return {
                    header: self.worksheet.cell(row=row_idx, column=col_idx).value or ""
                    for header, col_idx in headers.items()
                }

        raise ValueError(f"Test case '{test_case_no}' not found")

    def get_all_test_cases(self, sheet_name: str = "TestCases", execute_only: bool = True) -> List[Dict[str, Any]]:
        """Get all test cases from sheet"""
        headers, header_row = self._get_headers(sheet_name)
        test_cases: List[Dict[str, Any]] = []
        execute_col = headers.get('Execute')

        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=header_row + 1, values_only=False), header_row + 1):
            # Skip empty rows
            if not row[headers['TestCaseNo'] - 1].value:
                continue

            # Filter by Execute column if needed
            if execute_only:
                execute_value = self.worksheet.cell(row=row_idx, column=execute_col).value
                if str(execute_value).strip().lower() != 'yes':
                    continue

            test_cases.append({
                header: self.worksheet.cell(row=row_idx, column=col_idx).value or ""
                for header, col_idx in headers.items()
            })

        return test_cases

    def update_test_result(self, test_case_no: str, test_result: str, fail_description: str = "", sheet_name: str = "TestCases") -> bool:
        """Update TestResult and Fail_Description columns"""
        headers, header_row = self._get_headers(sheet_name)
        test_case_col = headers.get('TestCaseNo')
        test_result_col = headers.get('TestResult')
        fail_desc_col = headers.get('Fail_Description')

        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=header_row + 1, values_only=False), header_row + 1):
            if row[test_case_col - 1].value == test_case_no:
                if test_result_col:
                    self.worksheet.cell(row=row_idx, column=test_result_col).value = test_result

                if fail_desc_col and fail_description:
                    self.worksheet.cell(row=row_idx, column=fail_desc_col).value = fail_description

                result_folder_col = headers.get('ResultFolder')
                if result_folder_col:
                    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
                    self.worksheet.cell(row=row_idx, column=result_folder_col).value = f"results/{timestamp}"

                self.workbook.save(self.file_path)
                return True

        raise ValueError(f"Test case '{test_case_no}' not found")

    def get_test_input_data(self, test_case_no: str, sheet_name: str = "TestData") -> Dict[str, Any]:
        """Get test input data from TestData sheet"""
        return self.get_test_case_data(test_case_no, sheet_name)

    def get_expected_result(self, test_case_no: str, sheet_name: str = "ExpectedResults") -> Dict[str, Any]:
        """Get expected result from ExpectedResults sheet"""
        return self.get_test_case_data(test_case_no, sheet_name)

    def create_excel_template(self, file_path: str, sheet_names: Optional[List[str]] = None) -> str:
        """Create Excel template file with required sheets and headers"""
        if sheet_names is None:
            sheet_names = ["TestCases", "TestData", "ExpectedResults"]

        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        sheet_configs = {
            "TestCases": {
                "headers": ["TestCaseNo", "ResultFolder", "Execute", "Module/Feature", "TestCaseDescription",
                           "TestResult", "Fail_Description", "ExpectedResult", "Priority", "Status"],
                "widths": {'A': 15, 'B': 25, 'C': 10, 'D': 20, 'E': 50, 'F': 15, 'G': 50, 'H': 50, 'I': 10, 'J': 10}
            },
            "TestData": {
                "headers": ["TestCaseNo", "Program_Type", "ID_Card", "Phone_Number", "OTP1", "OTP2",
                           "Card_Selection", "User_Choice", "Occupation", "Reason", "Payment_Amount",
                           "Company_Name", "Position", "Work_Years", "Work_Months", "Monthly_Income",
                           "Work_Phone", "Extension", "Mobile_Phone", "Email"],
                "widths": {}
            },
            "ExpectedResults": {
                "headers": ["TestCaseNo", "Page_Title", "Element_To_Verify", "Expected_Value", "Verification_Type"],
                "widths": {}
            }
        }

        for sheet_name in sheet_names:
            if sheet_name not in sheet_configs:
                continue

            config = sheet_configs[sheet_name]
            ws = wb.create_sheet(sheet_name)
            ws.append(config["headers"])

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            for col, width in config.get("widths", {}).items():
                ws.column_dimensions[col].width = width

        wb.save(file_path)
        return file_path
