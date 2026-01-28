"""One-off script: set Card_Selection to 16-digit numeric in all test data xlsx files."""
import os
import tempfile
from pathlib import Path
from openpyxl import load_workbook

# Match current HTML: first4=9999, last4=4322
CARD_NUMBER = 9999432109874322
TEST_DATA_DIR = Path(__file__).parent / "test_data"
FILES = [
    "test_data.xlsx",
    "test_data_debt_re.xlsx",
    "test_data_errors.xlsx",
    "test_data_rewrite_settlement.xlsx",
]


def update_file(path: Path) -> bool:
    if not path.exists():
        print(f"Skip (not found): {path}")
        return False
    wb = load_workbook(path)
    if "TestData" not in wb.sheetnames:
        print(f"Skip (no TestData sheet): {path}")
        wb.close()
        return False
    ws = wb["TestData"]
    headers = {}
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if row and any(str(x) == "TestCaseNo" for x in row if x):
            for col_idx, val in enumerate(row, 1):
                if val:
                    headers[str(val).strip()] = col_idx
            header_row = row_idx
            break
    if "Card_Selection" not in headers or header_row is None:
        print(f"Skip (no Card_Selection col): {path}")
        wb.close()
        return False
    col = headers["Card_Selection"]
    test_case_col = headers.get("TestCaseNo", 1)
    count = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=test_case_col).value:
            ws.cell(row=row_idx, column=col).value = CARD_NUMBER
            count += 1
    try:
        wb.save(path)
        wb.close()
        print(f"Updated {count} row(s) in {path.name}")
        return True
    except PermissionError:
        fallback = path.parent / (path.stem + "_numeric.xlsx")
        try:
            wb.save(fallback)
            print(f"Permission denied on {path.name}. Wrote {fallback.name} — copy over {path.name} when closed.")
            print(f"Updated {count} row(s).")
            return True
        except Exception:
            print(f"Permission denied (close in Excel?): {path.name}")
            return False
        finally:
            wb.close()


if __name__ == "__main__":
    for name in FILES:
        update_file(TEST_DATA_DIR / name)
