"""
Script to populate sample test data in Excel files
"""
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

def populate_sample_data():
    """Populate sample test cases in test_data.xlsx"""
    project_root = Path(__file__).parent
    excel_file = project_root / "test_data" / "test_data.xlsx"
    
    wb = load_workbook(excel_file)
    
    # Populate TestCases sheet
    ws = wb["TestCases"]
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    result_folder = f"results/{timestamp}"
    
    test_cases = [
        ["TC_AUTH_001", result_folder, "Yes", "Authentication", "Valid ID card and phone number", "", "", "Navigate to OTP1 page", "High", "Active"],
        ["TC_AUTH_002", result_folder, "Yes", "Authentication", "Invalid ID card format", "", "", "Show error: คุณกรอกข้อมูลผิดรูปแบบ", "Medium", "Active"],
        ["TC_AUTH_005", result_folder, "Yes", "Authentication", "5 failed attempts - Account locked", "", "", "Account locked for 10 minutes", "High", "Active"],
        ["TC_OTP1_001", result_folder, "Yes", "OTP1", "Valid OTP (123456)", "", "", "Navigate to Select Credit Card", "High", "Active"],
        ["TC_OTP1_002", result_folder, "Yes", "OTP1", "Wrong OTP", "", "", "Show error, allow retry", "Medium", "Active"],
        ["TC_CARD_001", result_folder, "Yes", "Select Card", "Select eligible card", "", "", "Navigate to Offering Program", "High", "Active"],
        ["TC_PROG_001", result_folder, "Yes", "Offering Program", "Debt Re - Show 2 programs", "", "", "Display 2 programs, none selected", "High", "Active"],
        ["TC_PROG_002", result_folder, "Yes", "Offering Program", "Rewrite Settlement - Show 1 program with discount", "", "", "Display 1 program, selected, with discount", "High", "Active"],
        ["TC_DETAIL_001", result_folder, "Yes", "Program Details", "Debt Re - Verify title and no conditions", "", "", "Title: โครงการปรับปรุงโครงสร้างหนี้, No conditions", "High", "Active"],
        ["TC_DETAIL_002", result_folder, "Yes", "Program Details", "Rewrite Settlement - Verify title and conditions", "", "", "Title: โครงการปรับปรุงโครงสร้างหนี้แบบมีส่วนลด, Has conditions", "High", "Active"],
        ["TC_PAY_001", result_folder, "Yes", "Set Payment", "Enter amount and calculate", "", "", "Show calculation results", "Medium", "Active"],
        ["TC_RISK_001", result_folder, "Yes", "Risk Questionnaire", "Fill all fields - Private Employee", "", "", "Show review page with all data", "High", "Active"],
        ["TC_ECON_001", result_folder, "Yes", "E-Contract", "Accept contract", "", "", "Navigate to OTP Confirmation", "High", "Active"],
        ["TC_INT_001", result_folder, "Yes", "Integration", "Full flow - Debt Re", "", "", "Complete registration successfully", "High", "Active"],
        ["TC_INT_002", result_folder, "Yes", "Integration", "Full flow - Rewrite Settlement", "", "", "Complete registration successfully", "High", "Active"],
    ]
    
    for row_data in test_cases:
        ws.append(row_data)
    
    # Populate TestData sheet
    if "TestData" in wb.sheetnames:
        ws_data = wb["TestData"]
        test_data = [
            ["TC_AUTH_001", "debt re", "1-2345-67890-12-3", "082-999-9999", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["TC_OTP1_001", "debt re", "", "", "123456", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["TC_PROG_001", "debt re", "", "", "", "", "card1", "interested", "", "", "", "", "", "", "", "", "", "", "", ""],
            ["TC_INT_001", "debt re", "1-2345-67890-12-3", "082-999-9999", "123456", "123456", "card1", "interested", "private_employee", "other", "", "บริษัท ABC", "พนักงาน", "9", "6", "80000", "023456789", "123", "0898765432", "example@email.com"],
        ]
        for row_data in test_data:
            ws_data.append(row_data)
    
    # Populate ExpectedResults sheet
    if "ExpectedResults" in wb.sheetnames:
        ws_expected = wb["ExpectedResults"]
        expected_results = [
            ["TC_PROG_001", "Offering Program", "Program Count", "2 โครงการ", "Text"],
            ["TC_PROG_002", "Offering Program", "Program Count", "1 โครงการ", "Text"],
            ["TC_PROG_002", "Offering Program", "Max Discount", "83,666 บาท", "Text"],
            ["TC_DETAIL_001", "Program Details", "Title", "โครงการปรับปรุงโครงสร้างหนี้", "Text"],
            ["TC_DETAIL_002", "Program Details", "Payment Conditions", "Visible", "Visibility"],
            ["TC_DETAIL_002", "Program Details", "Contact Info", "02-777-1555", "Text"],
        ]
        for row_data in expected_results:
            ws_expected.append(row_data)
    
    wb.save(excel_file)
    print(f"Sample test data populated in {excel_file}")

if __name__ == "__main__":
    populate_sample_data()






